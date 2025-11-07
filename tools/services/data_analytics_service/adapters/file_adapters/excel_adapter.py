#!/usr/bin/env python3
"""
Excel file adapter for metadata extraction
"""

from typing import Dict, List, Any, Optional
from collections import Counter
from .base_adapter import FileAdapter
from .base_adapter import TableInfo, ColumnInfo

try:
    import polars as pl
    import openpyxl
    POLARS_AVAILABLE = True
    OPENPYXL_AVAILABLE = True
except ImportError:
    POLARS_AVAILABLE = False
    OPENPYXL_AVAILABLE = False

class ExcelAdapter(FileAdapter):
    """Excel file adapter for .xlsx and .xls files"""
    
    def __init__(self):
        super().__init__()
        if not POLARS_AVAILABLE:
            raise ImportError("polars is required for Excel adapter. Install with: pip install polars")
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel adapter. Install with: pip install openpyxl")
        
        self.workbook = None
        self.sheet_data = {}
        self.sheet_names = []
    
    def _load_file(self, file_path: str) -> Any:
        """Load Excel file and all sheets"""
        try:
            # Load workbook to get sheet names first
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            self.sheet_names = wb.sheetnames
            wb.close()

            # Load each sheet using Polars
            sheet_data = {}
            for sheet_name in self.sheet_names:
                try:
                    df = pl.read_excel(file_path, sheet_name=sheet_name)
                    sheet_data[sheet_name] = df
                except Exception as e:
                    print(f"Warning: Could not load sheet '{sheet_name}': {e}")
                    sheet_data[sheet_name] = pl.DataFrame()  # Empty dataframe
            
            self.sheet_data = sheet_data
            
            # Also load workbook for additional metadata
            try:
                self.workbook = openpyxl.load_workbook(file_path, data_only=True)
            except Exception as e:
                print(f"Warning: Could not load workbook metadata: {e}")
                self.workbook = None
            
            return sheet_data
        
        except Exception as e:
            raise Exception(f"Failed to load Excel file: {e}")
    
    def _analyze_structure(self) -> Dict[str, Any]:
        """Analyze Excel file structure"""
        structure = {
            "total_sheets": len(self.sheet_names),
            "sheet_names": self.sheet_names,
            "sheets_info": {}
        }
        
        for sheet_name, df in self.sheet_data.items():
            sheet_info = {
                "rows": df.height,
                "columns": df.width,
                "column_names": list(df.columns),
                "has_header": self._detect_header(df),
                "data_types": {str(col): str(df.get_column(str(col)).dtype) for col in df.columns},
                "empty_rows": df.select(pl.all().is_null()).sum_horizontal().sum(),
                "empty_columns": df.null_count().sum()
            }
            structure["sheets_info"][sheet_name] = sheet_info
        
        return structure
    
    def _detect_header(self, df: pl.DataFrame) -> bool:
        """Detect if first row is header"""
        if df.height == 0:
            return False
        
        # Check if first row has different data types than rest
        if df.height < 2:
            return True  # Assume header if only one row
        
        first_row = df.row(0)
        second_row = df.row(1)
        
        # If first row is all strings and second row has numbers, likely header
        first_all_strings = all(isinstance(val, str) for val in first_row if not pl.is_null(val))
        second_has_numbers = any(isinstance(val, (int, float)) for val in second_row if not pl.is_null(val))
        
        return first_all_strings and second_has_numbers
    
    def get_tables(self, schema_filter: Optional[List[str]] = None) -> List[TableInfo]:
        """Get sheet information as table metadata"""
        tables = []
        
        for sheet_name, df in self.sheet_data.items():
            # Skip sheets based on filter
            if schema_filter and sheet_name not in schema_filter:
                continue
            
            table_info = TableInfo(
                table_name=sheet_name,
                schema_name="excel_file",
                table_type="SHEET",
                record_count=df.height,
                table_comment=f"Excel sheet: {sheet_name}",
                created_date="",
                last_modified="",
                business_category=self._categorize_sheet(sheet_name, df),
                update_frequency="unknown",
                data_quality_score=self._calculate_data_quality_score(df),
                key_columns=self._identify_key_columns(df)
            )
            
            tables.append(table_info)
        
        return tables
    
    def get_columns(self, table_name: Optional[str] = None) -> List[ColumnInfo]:
        """Get column information from Excel sheets"""
        columns = []
        
        sheets_to_process = [table_name] if table_name else self.sheet_names
        
        for sheet_name in sheets_to_process:
            if sheet_name not in self.sheet_data:
                continue
            
            df = self.sheet_data[sheet_name]
            
            for idx, column_name in enumerate(df.columns):
                column_series = df.get_column(column_name)
                
                # Analyze column data
                data_analysis = self._analyze_column_data(column_series)
                
                column_info = ColumnInfo(
                    table_name=sheet_name,
                    column_name=str(column_name),
                    data_type=self._map_polars_to_sql_type(column_series.dtype),
                    max_length=data_analysis.get('max_length'),
                    is_nullable=column_series.null_count() > 0,
                    default_value="",
                    column_comment=f"Column from Excel sheet {sheet_name}",
                    ordinal_position=idx + 1,
                    business_type=self._infer_business_type(column_name, column_series),
                    value_range=data_analysis.get('value_range'),
                    null_percentage=data_analysis.get('null_percentage'),
                    unique_percentage=data_analysis.get('unique_percentage'),
                    sample_values=data_analysis.get('sample_values')
                )
                
                columns.append(column_info)
        
        return columns
    
    def _analyze_column_data(self, series: pl.Series) -> Dict[str, Any]:
        """Analyze polars series data"""
        analysis = {}
        
        # Basic stats
        total_count = series.len()
        null_count = series.null_count()
        non_null_series = series.drop_nulls()
        
        analysis['null_percentage'] = null_count / total_count if total_count > 0 else 0
        analysis['unique_count'] = non_null_series.n_unique()
        analysis['unique_percentage'] = analysis['unique_count'] / non_null_series.len() if non_null_series.len() > 0 else 0
        
        # Sample values
        sample_size = min(10, non_null_series.len())
        if sample_size > 0:
            analysis['sample_values'] = non_null_series.sample(n=sample_size).to_list()
        else:
            analysis['sample_values'] = []
        
        # Value range for numeric columns
        if series.dtype in [pl.Int8, pl.Int16, pl.Int32, pl.Int64, pl.UInt8, pl.UInt16, pl.UInt32, pl.UInt64, pl.Float32, pl.Float64]:
            if non_null_series.len() > 0:
                analysis['value_range'] = {
                    'min': float(non_null_series.min()),
                    'max': float(non_null_series.max()),
                    'mean': float(non_null_series.mean())
                }
        
        # Max length for string columns
        if series.dtype == pl.Utf8 or series.dtype == pl.Object:
            string_lengths = non_null_series.cast(pl.Utf8).str.len_chars()
            if string_lengths.len() > 0:
                analysis['max_length'] = int(string_lengths.max())
            else:
                analysis['max_length'] = 0
        
        return analysis
    
    def _map_polars_to_sql_type(self, dtype) -> str:
        """Map polars dtype to SQL-like type"""
        dtype_str = str(dtype)
        
        if 'int' in dtype_str:
            return 'integer'
        elif 'float' in dtype_str:
            return 'decimal'
        elif 'bool' in dtype_str:
            return 'boolean'
        elif 'datetime' in dtype_str:
            return 'timestamp'
        elif 'object' in dtype_str:
            return 'text'
        else:
            return 'text'
    
    def _infer_business_type(self, column_name: str, series: pl.Series) -> str:
        """Infer business type from column name and data"""
        column_lower = str(column_name).lower()
        
        # ID columns
        if any(keyword in column_lower for keyword in ['id', 'key', 'code', 'number']):
            return 'identifier'
        
        # Name columns
        elif any(keyword in column_lower for keyword in ['name', 'title', 'description']):
            return 'name'
        
        # Amount/Price columns
        elif any(keyword in column_lower for keyword in ['amount', 'price', 'cost', 'value', 'total']):
            return 'amount'
        
        # Date columns
        elif any(keyword in column_lower for keyword in ['date', 'time', 'created', 'updated']):
            return 'datetime'
        
        # Status columns
        elif any(keyword in column_lower for keyword in ['status', 'state', 'type', 'category']):
            return 'category'
        
        # Email columns
        elif 'email' in column_lower or '@' in str(series[0] if series.len() > 0 else ''):
            return 'email'
        
        # Phone columns
        elif any(keyword in column_lower for keyword in ['phone', 'tel', 'mobile']):
            return 'phone'
        
        else:
            return 'general'
    
    def _categorize_sheet(self, sheet_name: str, df: pl.DataFrame) -> str:
        """Categorize sheet based on name and content"""
        sheet_lower = sheet_name.lower()
        
        if any(keyword in sheet_lower for keyword in ['config', 'setting', 'parameter']):
            return 'configuration'
        elif any(keyword in sheet_lower for keyword in ['lookup', 'reference', 'master']):
            return 'reference'
        elif any(keyword in sheet_lower for keyword in ['transaction', 'detail', 'record']):
            return 'transaction'
        elif any(keyword in sheet_lower for keyword in ['summary', 'report', 'dashboard']):
            return 'report'
        else:
            return 'data'
    
    def _calculate_data_quality_score(self, df: pl.DataFrame) -> float:
        """Calculate data quality score for sheet"""
        if df.height == 0:
            return 0.0
        
        total_cells = df.height * df.width
        null_cells = df.null_count().sum_horizontal()[0]
        completeness = 1 - (null_cells / total_cells)
        
        # Simple quality score based on completeness
        return round(completeness, 2)
    
    def _identify_key_columns(self, df: pl.DataFrame) -> List[str]:
        """Identify potential key columns"""
        key_columns = []
        
        for column in df.columns:
            column_lower = str(column).lower()
            series = df.get_column(column)
            
            # Check if column name suggests it's a key
            if any(keyword in column_lower for keyword in ['id', 'key', 'code']):
                key_columns.append(str(column))
            
            # Check if column has unique values (potential key)
            elif series.n_unique() == len(series.drop_nulls()) and len(series.drop_nulls()) > 0:
                key_columns.append(str(column))
        
        return key_columns[:3]  # Return max 3 key columns
    
    def _get_sample_data_internal(self, sheet_name: str, limit: int) -> List[Dict[str, Any]]:
        """Get sample data from Excel sheet"""
        if sheet_name not in self.sheet_data:
            return []
        
        df = self.sheet_data[sheet_name]
        sample_df = df.head(limit)
        
        # Convert to list of dictionaries
        return sample_df.to_dicts()
    
    def _analyze_column_distribution(self, sheet_name: str, column_name: str, sample_size: int) -> Dict[str, Any]:
        """Analyze column data distribution"""
        if sheet_name not in self.sheet_data:
            return {"error": f"Sheet {sheet_name} not found"}
        
        df = self.sheet_data[sheet_name]
        
        if column_name not in df.columns:
            return {"error": f"Column {column_name} not found in sheet {sheet_name}"}
        
        series = df.get_column(column_name)
        return self._analyze_column_data(series)
    
    def _compute_quality_metrics(self, metrics: Dict[str, Any]) -> Dict[str, Any]:
        """Compute Excel-specific quality metrics"""
        metrics['total_sections'] = len(self.sheet_names)
        
        total_columns = 0
        total_rows = 0
        empty_sheets = 0
        columns_with_nulls = 0
        
        for sheet_name, df in self.sheet_data.items():
            if df.height == 0:
                empty_sheets += 1
                continue
            
            total_columns += df.width
            total_rows += df.height
            
            # Count columns with null values
            for column in df.columns:
                if df.get_column(column).is_null().any():
                    columns_with_nulls += 1
        
        metrics.update({
            'total_columns': total_columns,
            'total_rows': total_rows,
            'empty_sections': empty_sheets,
            'columns_with_nulls': columns_with_nulls
        })
        
        return metrics
    
    def _get_data_summary(self) -> Dict[str, Any]:
        """Get summary of Excel data"""
        summary = {
            'total_sheets': len(self.sheet_names),
            'sheet_summaries': {}
        }
        
        for sheet_name, df in self.sheet_data.items():
            sheet_summary = {
                'rows': df.height,
                'columns': df.width,
                'data_types': df.dtypes,
                'memory_usage': df.estimated_size("b")
            }
            summary['sheet_summaries'][sheet_name] = sheet_summary
        
        return summary
    
    def get_workbook_metadata(self) -> Dict[str, Any]:
        """Get Excel workbook metadata"""
        metadata = {
            'file_info': self.file_info,
            'sheets': {}
        }
        
        if self.workbook:
            for sheet_name in self.workbook.sheetnames:
                worksheet = self.workbook[sheet_name]
                sheet_metadata = {
                    'title': worksheet.title,
                    'max_row': worksheet.max_row,
                    'max_column': worksheet.max_column,
                    'sheet_state': worksheet.sheet_state,
                    'sheet_properties': {
                        'tab_color': str(worksheet.sheet_properties.tabColor) if worksheet.sheet_properties.tabColor else None
                    }
                }
                metadata['sheets'][sheet_name] = sheet_metadata
        
        return metadata
    
    def export_sheet_to_csv(self, sheet_name: str, output_path: str) -> bool:
        """Export specific sheet to CSV"""
        try:
            if sheet_name not in self.sheet_data:
                return False
            
            df = self.sheet_data[sheet_name]
            df.write_csv(output_path)
            return True
        except Exception:
            return False